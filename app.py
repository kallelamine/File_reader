"""
Flask web application for extracting structured data from Tunisian Ministry of Finance documents
using OpenAI Vision API and generating XLSX files.
"""

import os
import json
import uuid
import re
import base64
from datetime import datetime
from pathlib import Path
from werkzeug.utils import secure_filename
from flask import Flask, request, render_template, send_file, jsonify, flash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from openai import OpenAI
import zipfile
import io

# Load environment variables
load_dotenv()

# Vercel serverless: only /tmp is writable
IS_VERCEL = os.environ.get('VERCEL') == '1'
if IS_VERCEL:
    UPLOAD_DIR = '/tmp/uploads'
    OUTPUT_DIR = '/tmp/outputs'
else:
    UPLOAD_DIR = 'uploads'
    OUTPUT_DIR = 'outputs'

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
app.config['OUTPUT_FOLDER'] = OUTPUT_DIR

# Allowed file extensions
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'PNG', 'JPG', 'JPEG'}

# OpenAI client: lazy init so app loads even if OPEN_AI_API_KEY is missing (e.g. Vercel env not set)
_openai_client = None

def get_openai_client():
    """Return OpenAI client; raise only when actually used without API key."""
    global _openai_client
    if _openai_client is not None:
        return _openai_client
    key = os.getenv('OPEN_AI_API_KEY')
    if not key:
        raise ValueError("OPEN_AI_API_KEY not found. Set it in Vercel Project Settings → Environment Variables.")
    _openai_client = OpenAI(api_key=key)
    return _openai_client

# Create necessary directories (/tmp on Vercel is writable)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)


def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def normalize_numeric(value):
    """Normalize numeric values by removing spaces and converting to number."""
    if not value or value == '':
        return ''
    if isinstance(value, (int, float)):
        return value
    # Remove spaces and convert to number if possible
    cleaned = str(value).replace(' ', '').replace(',', '.')
    try:
        # Try to convert to float first, then int if it's a whole number
        num = float(cleaned)
        if num.is_integer():
            return int(num)
        return num
    except (ValueError, AttributeError):
        return value


def extract_with_openai(image_path):
    """
    Send image to OpenAI Vision API and extract structured JSON.
    Returns parsed JSON or None if extraction fails.
    """
    try:
        # Read and encode image to base64
        with open(image_path, 'rb') as image_file:
            image_data = image_file.read()
            base64_image = base64.b64encode(image_data).decode('utf-8')
            
            # Determine image MIME type from file extension
            ext = Path(image_path).suffix.lower()
            if ext in ['.png']:
                mime_type = 'image/png'
            else:
                mime_type = 'image/jpeg'
            
            response = get_openai_client().chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert document parser for Tunisian Ministry of Finance registry documents."
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": """This image contains financial registry data titled either:

- "Actes sur les Sociétés"
or
- "BIENS IMMOBILIERS (Qualité Acheteur)"

Return STRICT JSON ONLY.

First determine:

doc_type:
- ACTES_SOCIETES
- BIENS_IMMOBILIERS_ACHETEUR

Then extract:

COMMON HEADER:

edited_date  
page_info  
matricule_fiscal  
cin_number  
raison_sociale  
activite  
adresse  

---

If doc_type == ACTES_SOCIETES:

For each row/year:

annee  
ref_enregistrement  
date_enregistrement  
type_acte  
date_acte  
matricule_fiscal_societe  
raison_sociale_societe  
capital_societe  
forme_juridique  
apport_numeraire  
apport_nature  
apport_fonds_commerce  
apport_incorporation  
apport_creances  
apport_autres  
total_apports  

Also:

total_annuel  
total_general  

Return array under key: actes_societes

---

If doc_type == BIENS_IMMOBILIERS_ACHETEUR:

This is a structured table where:
- Each horizontal line represents ONE property transaction
- The year ("Année: 2010") applies to ALL rows until the next year header
- Vendor info belongs to SAME row as property
- Property info belongs to SAME row

HEADER EXTRACTION:
edited_date
page_info
matricule_fiscal
raison_sociale
activite
adresse
annee_courante

ROW EXTRACTION (MANDATORY PER ROW):

For EACH property row extract EXACTLY:

annee                         (from annee_courante, applies to all rows until next year)
ref_enregistrement
date_enregistrement
numero_quittance
date_quittance
type_acte
nature_acte
date_acte
nbr_parts

vendeur_matricule_fiscal
vendeur_cin
vendeur_nom

numero_bien
nature_et_adresse_bien        (FULL text, do not truncate)
recette_et_date_origine
surface_bien                   (numeric if visible)
montant_vente_bien            (capture exactly, preserve original AND normalized)

TOTAL:
At end of year block capture:
total_annuel

RULES:
1. Never merge multiple properties into one row.
2. Vendor info belongs to SAME row.
3. Property info belongs to SAME row.
4. Nature + Adresse du bien must be FULL text.
5. Surface must be numeric if visible.
6. Montant vente must be captured exactly.
7. If field is empty → return "".
8. Preserve original number text AND also provide normalized version if confident.

Return array under key: biens_immobiliers

---

JSON FORMAT:

{
  "doc_type": "",
  "header": {
    "edited_date": "",
    "page_info": "",
    "matricule_fiscal": "",
    "cin_number": "",
    "raison_sociale": "",
    "activite": "",
    "adresse": "",
    "annee_courante": ""
  },
  "actes_societes": [],
  "biens_immobiliers": [
    {
      "annee": "",
      "ref_enregistrement": "",
      "date_enregistrement": "",
      "numero_quittance": "",
      "date_quittance": "",
      "type_acte": "",
      "nature_acte": "",
      "date_acte": "",
      "nbr_parts": "",
      "vendeur_matricule_fiscal": "",
      "vendeur_cin": "",
      "vendeur_nom": "",
      "numero_bien": "",
      "nature_et_adresse_bien": "",
      "recette_et_date_origine": "",
      "surface_bien": "",
      "montant_vente_bien": "",
      "total_annuel": ""
    }
  ]
}

No markdown. No explanations.

Use empty string if field not found."""
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mime_type};base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                response_format={"type": "json_object"},
                temperature=0.1
            )
            
            # Parse JSON response
            content = response.choices[0].message.content
            return json.loads(content)
    except Exception as e:
        print(f"OpenAI extraction error: {str(e)}")
        return None


def create_mock_data(doc_type):
    """Create mock data structure if OpenAI extraction fails."""
    if doc_type == "ACTES_SOCIETES":
        return {
            "doc_type": "ACTES_SOCIETES",
            "header": {
                "edited_date": "",
                "page_info": "",
                "matricule_fiscal": "",
                "cin_number": "",
                "raison_sociale": "",
                "activite": "",
                "adresse": ""
            },
            "actes_societes": [],
            "biens_immobiliers": []
        }
    else:
        return {
            "doc_type": "BIENS_IMMOBILIERS_ACHETEUR",
            "header": {
                "edited_date": "",
                "page_info": "",
                "matricule_fiscal": "",
                "cin_number": "",
                "raison_sociale": "",
                "activite": "",
                "adresse": ""
            },
            "actes_societes": [],
            "biens_immobiliers": []
        }


def create_xlsx_actes_societes(data, output_path, photo_name):
    """Create XLSX file for ACTES_SOCIETES document type."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    
    # Metadata rows - include all header fields
    header_data = data.get("header", {})
    metadata = [
        ["photo_name", photo_name],
        ["doc_type", data.get("doc_type", "")],
        ["processed_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["edited_date", header_data.get("edited_date", "")],
        ["page_info", header_data.get("page_info", "")],
        ["matricule_fiscal", header_data.get("matricule_fiscal", "")],
        ["cin_number", header_data.get("cin_number", "")],
        ["raison_sociale", header_data.get("raison_sociale", "")],
        ["activite", header_data.get("activite", "")],
        ["adresse", header_data.get("adresse", "")],
        ["annee_courante", header_data.get("annee_courante", "")]
    ]
    
    for idx, (key, value) in enumerate(metadata, start=1):
        ws.cell(row=idx, column=1, value=key)
        ws.cell(row=idx, column=2, value=value)
    
    # Column headers (row after metadata)
    header_row = len(metadata) + 1
    headers = [
        "annee", "ref_enregistrement", "date_enregistrement", "type_acte",
        "date_acte", "matricule_fiscal_societe", "raison_sociale_societe",
        "capital_societe", "forme_juridique", "apport_numeraire",
        "apport_nature", "apport_fonds_commerce", "apport_incorporation",
        "apport_creances", "apport_autres", "total_apports",
        "total_annuel", "total_general"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    actes = data.get("actes_societes", [])
    for row_idx, acte in enumerate(actes, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=acte.get("annee", ""))
        ws.cell(row=row_idx, column=2, value=acte.get("ref_enregistrement", ""))
        ws.cell(row=row_idx, column=3, value=acte.get("date_enregistrement", ""))
        ws.cell(row=row_idx, column=4, value=acte.get("type_acte", ""))
        ws.cell(row=row_idx, column=5, value=acte.get("date_acte", ""))
        ws.cell(row=row_idx, column=6, value=acte.get("matricule_fiscal_societe", ""))
        ws.cell(row=row_idx, column=7, value=acte.get("raison_sociale_societe", ""))
        ws.cell(row=row_idx, column=8, value=normalize_numeric(acte.get("capital_societe", "")))
        ws.cell(row=row_idx, column=9, value=acte.get("forme_juridique", ""))
        ws.cell(row=row_idx, column=10, value=normalize_numeric(acte.get("apport_numeraire", "")))
        ws.cell(row=row_idx, column=11, value=normalize_numeric(acte.get("apport_nature", "")))
        ws.cell(row=row_idx, column=12, value=normalize_numeric(acte.get("apport_fonds_commerce", "")))
        ws.cell(row=row_idx, column=13, value=normalize_numeric(acte.get("apport_incorporation", "")))
        ws.cell(row=row_idx, column=14, value=normalize_numeric(acte.get("apport_creances", "")))
        ws.cell(row=row_idx, column=15, value=normalize_numeric(acte.get("apport_autres", "")))
        ws.cell(row=row_idx, column=16, value=normalize_numeric(acte.get("total_apports", "")))
        ws.cell(row=row_idx, column=17, value=normalize_numeric(acte.get("total_annuel", "")))
        ws.cell(row=row_idx, column=18, value=normalize_numeric(acte.get("total_general", "")))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def create_xlsx_biens_immobiliers(data, output_path, photo_name):
    """Create XLSX file for BIENS_IMMOBILIERS_ACHETEUR document type."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    
    # Metadata rows - include all header fields
    header_data = data.get("header", {})
    metadata = [
        ["photo_name", photo_name],
        ["doc_type", data.get("doc_type", "")],
        ["processed_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["edited_date", header_data.get("edited_date", "")],
        ["page_info", header_data.get("page_info", "")],
        ["matricule_fiscal", header_data.get("matricule_fiscal", "")],
        ["cin_number", header_data.get("cin_number", "")],
        ["raison_sociale", header_data.get("raison_sociale", "")],
        ["activite", header_data.get("activite", "")],
        ["adresse", header_data.get("adresse", "")],
        ["annee_courante", header_data.get("annee_courante", "")]
    ]
    
    for idx, (key, value) in enumerate(metadata, start=1):
        ws.cell(row=idx, column=1, value=key)
        ws.cell(row=idx, column=2, value=value)
    
    # Column headers (row after metadata)
    header_row = len(metadata) + 1
    headers = [
        "annee", "ref_enregistrement", "date_enregistrement", "numero_quittance",
        "date_quittance", "type_acte", "nature_acte", "date_acte", "nbr_parts",
        "vendeur_matricule_fiscal", "vendeur_cin", "vendeur_nom",
        "numero_bien", "nature_et_adresse_bien", "recette_et_date_origine",
        "surface_bien", "montant_vente_bien", "total_annuel"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    biens = data.get("biens_immobiliers", [])
    current_annee = ""
    current_total_annuel = ""
    
    for row_idx, bien in enumerate(biens, start=header_row + 1):
        # Update current year if present in row
        if bien.get("annee"):
            current_annee = bien.get("annee", "")
        if bien.get("total_annuel"):
            current_total_annuel = bien.get("total_annuel", "")
        
        ws.cell(row=row_idx, column=1, value=current_annee if current_annee else bien.get("annee", ""))
        ws.cell(row=row_idx, column=2, value=bien.get("ref_enregistrement", ""))
        ws.cell(row=row_idx, column=3, value=bien.get("date_enregistrement", ""))
        ws.cell(row=row_idx, column=4, value=bien.get("numero_quittance", ""))
        ws.cell(row=row_idx, column=5, value=bien.get("date_quittance", ""))
        ws.cell(row=row_idx, column=6, value=bien.get("type_acte", ""))
        ws.cell(row=row_idx, column=7, value=bien.get("nature_acte", ""))
        ws.cell(row=row_idx, column=8, value=bien.get("date_acte", ""))
        ws.cell(row=row_idx, column=9, value=normalize_numeric(bien.get("nbr_parts", "")))
        ws.cell(row=row_idx, column=10, value=bien.get("vendeur_matricule_fiscal", ""))
        ws.cell(row=row_idx, column=11, value=bien.get("vendeur_cin", ""))
        ws.cell(row=row_idx, column=12, value=bien.get("vendeur_nom", ""))
        ws.cell(row=row_idx, column=13, value=bien.get("numero_bien", ""))
        ws.cell(row=row_idx, column=14, value=bien.get("nature_et_adresse_bien", ""))
        ws.cell(row=row_idx, column=15, value=bien.get("recette_et_date_origine", ""))
        ws.cell(row=row_idx, column=16, value=normalize_numeric(bien.get("surface_bien", "")))
        ws.cell(row=row_idx, column=17, value=normalize_numeric(bien.get("montant_vente_bien", "")))
        ws.cell(row=row_idx, column=18, value=normalize_numeric(current_total_annuel if current_total_annuel else bien.get("total_annuel", "")))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def extract_bilan_with_openai(image_path):
    """
    Extract BILAN table from an image containing "2-3 Etat de résultat et bilan de l'entreprise : BILAN".
    Returns parsed JSON with bilan data or None if extraction fails.
    """
    try:
        # Read and encode image to base64
        with open(image_path, 'rb') as image_file:
            image_data = image_file.read()
            base64_image = base64.b64encode(image_data).decode('utf-8')
            
            # Determine image MIME type from file extension
            ext = Path(image_path).suffix.lower()
            if ext in ['.png']:
                mime_type = 'image/png'
            else:
                mime_type = 'image/jpeg'
            
            response = get_openai_client().chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert extractor for 'Investigation Économique – Vision 360' documents (Tunisia)."
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": """IE VISION 360 BILAN EXTRACTION:
You are extracting the "BILAN – Etat de résultat" table from a Tunisian Investigation Economique (Vision 360) document.

Return STRICT JSON ONLY.

This BILAN is ALWAYS a MATRIX:

- One header row with EXACTLY 3 dates (left → right)
- Six KPI rows
- Each KPI has EXACTLY 3 values aligned to those dates

--------------------------------
STEP 1 — Detect the 3 dates
--------------------------------

Find the header line containing dates (example):

31/12/2022 31/12/2023 31/12/2024

Return them in left-to-right order as:

bilan_dates: ["31/12/2022","31/12/2023","31/12/2024"]

--------------------------------
STEP 2 — Extract KPI rows
--------------------------------

You MUST extract ALL SIX rows (even if empty):

1) Achats de l'année
2) Stock Initial
3) Stock Final
4) Résultat Fiscal
5) C.A Local Hors Taxes
6) C.A Total (T.T.C)

OCR variations allowed:
- "Achats de L'année"
- "CA Local HT"
- "CA Total TTC"

For EACH KPI return:

values_text: [v1, v2, v3]
values_norm: [n1, n2, n3]

Where:
- values_text keeps original spacing ("13 861 221")
- values_norm removes spaces ("13861221")

If unsure → values_norm = ""

--------------------------------
JSON FORMAT (STRICT)
--------------------------------

{
 "bilan": {
   "dates": ["","",""],
   "rows": {
     "achats_annee": {
       "values_text": ["","",""],
       "values_norm": ["","",""]
     },
     "stock_initial": {
       "values_text": ["","",""],
       "values_norm": ["","",""]
     },
     "stock_final": {
       "values_text": ["","",""],
       "values_norm": ["","",""]
     },
     "resultat_fiscal": {
       "values_text": ["","",""],
       "values_norm": ["","",""]
     },
     "ca_local_ht": {
       "values_text": ["","",""],
       "values_norm": ["","",""]
     },
     "ca_total_ttc": {
       "values_text": ["","",""],
       "values_norm": ["","",""]
     }
   }
 }
}

--------------------------------
STRICT RULES
--------------------------------

1. NEVER collapse to single values.
2. ALWAYS return arrays of length 3.
3. Preserve left-to-right alignment (value[0] aligns with date[0], value[1] with date[1], value[2] with date[2]).
4. Missing cell = "".
5. No markdown.
6. No explanation.
7. JSON only."""
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mime_type};base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                response_format={"type": "json_object"},
                temperature=0.1
            )
            
            # Parse JSON response
            content = response.choices[0].message.content
            return json.loads(content)
    except Exception as e:
        print(f"OpenAI BILAN extraction error: {str(e)}")
        return None


def create_ie_xlsx(data, output_path, photo_name, bilan_data=None):
    """
    Create IE_VISION_360 XLSX file with 6 sheets:
    1. requete
    2. identite_entreprise
    3. immobilisations_corporelles (if BIENS detected)
    4. immobilisations_financieres (if ACTES detected)
    5. bilan (always empty)
    6. metadata
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    header = data.get("header", {})
    actes = data.get("actes_societes", [])
    biens = data.get("biens_immobiliers", [])
    doc_type = data.get("doc_type", "")
    
    # Sheet 1: requete
    ws_requete = wb.create_sheet("requete")
    ws_requete.append(["date_demande", "raison_sociale", "matricule_raison_sociale", "date_livraison"])
    ws_requete.append([
        "",  # date_demande - empty
        header.get("raison_sociale", ""),
        header.get("matricule_fiscal", ""),
        ""   # date_livraison - empty
    ])
    
    # Style header row
    for cell in ws_requete[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 2: identite_entreprise
    ws_identite = wb.create_sheet("identite_entreprise")
    ws_identite.append([
        "denomination_commerciale", "adresse", "rne", "matricule_fiscale",
        "qualite", "forme_juridique", "statut_activite", "date_creation",
        "capital_societe", "representant_legal", "objet"
    ])
    
    # Get forme_juridique from first acte if available
    forme_juridique = ""
    capital_societe = ""
    if actes and len(actes) > 0:
        forme_juridique = actes[0].get("forme_juridique", "")
        capital_societe = normalize_numeric(actes[0].get("capital_societe", ""))
    
    ws_identite.append([
        header.get("raison_sociale", ""),  # denomination_commerciale
        header.get("adresse", ""),         # adresse
        "",                                # rne - empty
        header.get("matricule_fiscal", ""), # matricule_fiscale
        "",                                # qualite - empty
        forme_juridique,                   # forme_juridique
        "",                                # statut_activite - empty
        "",                                # date_creation - empty
        capital_societe,                   # capital_societe
        "",                                # representant_legal - empty
        header.get("activite", "")         # objet
    ])
    
    # Style header row
    for cell in ws_identite[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 3: immobilisations_corporelles (only if BIENS detected)
    ws_immob_corp = wb.create_sheet("immobilisations_corporelles")
    ws_immob_corp.append([
        "annee", "ref_enregistrement", "date_enregistrement", "numero_quittance",
        "date_quittance", "type_acte", "nature_acte", "date_acte", "nbr_parts",
        "vendeur_nom", "vendeur_cin", "vendeur_matricule_fiscal",
        "numero_bien", "nature_et_adresse_bien", "recette_et_date_origine",
        "surface_bien", "montant_vente_bien"
    ])
    
    # Style header row
    for cell in ws_immob_corp[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add biens data if available
    current_annee_ie = ""
    for bien in biens:
        # Update current year if present in row
        if bien.get("annee"):
            current_annee_ie = bien.get("annee", "")
        
        ws_immob_corp.append([
            current_annee_ie if current_annee_ie else bien.get("annee", ""),
            bien.get("ref_enregistrement", ""),
            bien.get("date_enregistrement", ""),
            bien.get("numero_quittance", ""),
            bien.get("date_quittance", ""),
            bien.get("type_acte", ""),
            bien.get("nature_acte", ""),
            bien.get("date_acte", ""),
            normalize_numeric(bien.get("nbr_parts", "")),
            bien.get("vendeur_nom", ""),
            bien.get("vendeur_cin", ""),
            bien.get("vendeur_matricule_fiscal", ""),
            bien.get("numero_bien", ""),
            bien.get("nature_et_adresse_bien", ""),
            bien.get("recette_et_date_origine", ""),
            normalize_numeric(bien.get("surface_bien", "")),
            normalize_numeric(bien.get("montant_vente_bien", ""))
        ])
    
    # Sheet 4: immobilisations_financieres (only if ACTES detected)
    ws_immob_fin = wb.create_sheet("immobilisations_financieres")
    ws_immob_fin.append([
        "annee", "ref_enregistrement", "date_enregistrement", "type_acte",
        "date_acte", "raison_sociale_societe", "capital_societe", "forme_juridique",
        "apport_numeraire", "apport_nature", "apport_fonds_commerce",
        "apport_incorporation", "apport_creances", "apport_autres", "total_apports"
    ])
    
    # Style header row
    for cell in ws_immob_fin[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add actes data if available
    for acte in actes:
        ws_immob_fin.append([
            acte.get("annee", ""),
            acte.get("ref_enregistrement", ""),
            acte.get("date_enregistrement", ""),
            acte.get("type_acte", ""),
            acte.get("date_acte", ""),
            acte.get("raison_sociale_societe", ""),
            normalize_numeric(acte.get("capital_societe", "")),
            acte.get("forme_juridique", ""),
            normalize_numeric(acte.get("apport_numeraire", "")),
            normalize_numeric(acte.get("apport_nature", "")),
            normalize_numeric(acte.get("apport_fonds_commerce", "")),
            normalize_numeric(acte.get("apport_incorporation", "")),
            normalize_numeric(acte.get("apport_creances", "")),
            normalize_numeric(acte.get("apport_autres", "")),
            normalize_numeric(acte.get("total_apports", ""))
        ])
    
    # Sheet 5: bilan
    ws_bilan = wb.create_sheet("bilan")
    
    # Extract BILAN data if available
    if bilan_data and "bilan" in bilan_data:
        bilan = bilan_data["bilan"]
        dates = bilan.get("dates", [])
        rows_data = bilan.get("rows", {})
        
        # Ensure we have exactly 3 dates (pad with empty strings if needed)
        while len(dates) < 3:
            dates.append("")
        dates = dates[:3]  # Take only first 3 dates
        
        # Create header row with dates (matrix format: row label + 3 date columns)
        header_row = [""]  # Empty first cell for row labels
        header_row.extend(dates)
        ws_bilan.append(header_row)
        
        # Style header row
        for cell in ws_bilan[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows - ALWAYS include all 6 KPIs (matrix format: one row × three columns)
        row_labels = [
            ("achats_annee", "Achats de l'année"),
            ("stock_initial", "Stock Initial"),
            ("stock_final", "Stock Final"),
            ("resultat_fiscal", "Résultat Fiscal"),
            ("ca_local_ht", "C.A Local Hors Taxes"),
            ("ca_total_ttc", "C.A Total (T.T.C)")
        ]
        
        for row_key, row_label in row_labels:
            row_data = rows_data.get(row_key, {})
            values_text = row_data.get("values_text", ["", "", ""])
            values_norm = row_data.get("values_norm", ["", "", ""])
            
            # Ensure arrays are exactly length 3
            while len(values_text) < 3:
                values_text.append("")
            while len(values_norm) < 3:
                values_norm.append("")
            values_text = values_text[:3]
            values_norm = values_norm[:3]
            
            # Use normalized values if available, otherwise use text values
            # Ensure proper alignment: value[0] with date[0], value[1] with date[1], value[2] with date[2]
            values = []
            for i in range(3):  # Exactly 3 columns
                if i < len(values_norm) and values_norm[i] and values_norm[i] != "":
                    # Prefer normalized numeric value
                    values.append(normalize_numeric(values_norm[i]) if values_norm[i] else "")
                elif i < len(values_text) and values_text[i] and values_text[i] != "":
                    # Fallback to text value
                    values.append(values_text[i])
                else:
                    values.append("")
            
            # Create matrix row: [row_label, value_col1, value_col2, value_col3]
            data_row = [row_label]
            data_row.extend(values)
            ws_bilan.append(data_row)
    else:
        # No BILAN data - create empty matrix structure with placeholder dates
        # Header row: empty label + 3 date columns
        ws_bilan.append(["", "Date 1", "Date 2", "Date 3"])
        
        # Style header row
        for cell in ws_bilan[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add all 6 KPI rows with empty values (matrix format)
        row_labels = [
            "Achats de l'année",
            "Stock Initial",
            "Stock Final",
            "Résultat Fiscal",
            "C.A Local Hors Taxes",
            "C.A Total (T.T.C)"
        ]
        
        for row_label in row_labels:
            ws_bilan.append([row_label, "", "", ""])  # One row × three empty columns
    
    # Sheet 6: metadata
    ws_metadata = wb.create_sheet("metadata")
    ws_metadata.append(["photo_name", photo_name])
    ws_metadata.append(["processed_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_metadata.append(["doc_type", doc_type])
    ws_metadata.append(["edited_date", header.get("edited_date", "")])
    ws_metadata.append(["page_info", header.get("page_info", "")])
    
    # Style metadata
    for row in ws_metadata.iter_rows():
        row[0].font = Font(bold=True)
    
    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def process_image(image_path, photo_name, batch_folder):
    """
    Process a single image: extract data and generate XLSX file(s).
    Returns list of generated XLSX filenames.
    """
    generated_files = []
    
    # Extract data using OpenAI
    extracted_data = extract_with_openai(image_path)
    
    if not extracted_data:
        # Fallback: try to determine doc_type from filename or use mock
        # For now, create empty files for both types as fallback
        print(f"Warning: OpenAI extraction failed for {photo_name}, creating empty files")
        extracted_data = {
            "doc_type": "UNKNOWN",
            "header": {},
            "actes_societes": [],
            "biens_immobiliers": []
        }
    
    # Get base filename without extension
    base_name = Path(photo_name).stem
    
    # Check if both document types are present
    has_actes = extracted_data.get("actes_societes") and len(extracted_data.get("actes_societes", [])) > 0
    has_biens = extracted_data.get("biens_immobiliers") and len(extracted_data.get("biens_immobiliers", [])) > 0
    doc_type = extracted_data.get("doc_type", "")
    
    # Determine which files to generate
    if has_actes and has_biens:
        # Both types detected - generate two files
        # ACTES_SOCIETES file
        actes_data = extracted_data.copy()
        actes_data["doc_type"] = "ACTES_SOCIETES"
        actes_data["biens_immobiliers"] = []
        actes_filename = f"{base_name}__ACTES_SOCIETES.xlsx"
        actes_path = os.path.join(batch_folder, actes_filename)
        create_xlsx_actes_societes(actes_data, actes_path, photo_name)
        generated_files.append(actes_filename)
        
        # BIENS_IMMOBILIERS file
        biens_data = extracted_data.copy()
        biens_data["doc_type"] = "BIENS_IMMOBILIERS_ACHETEUR"
        biens_data["actes_societes"] = []
        biens_filename = f"{base_name}__BIENS_IMMO.xlsx"
        biens_path = os.path.join(batch_folder, biens_filename)
        create_xlsx_biens_immobiliers(biens_data, biens_path, photo_name)
        generated_files.append(biens_filename)
        
    elif doc_type == "ACTES_SOCIETES" or has_actes:
        # Only ACTES_SOCIETES
        actes_data = extracted_data.copy()
        actes_data["doc_type"] = "ACTES_SOCIETES"
        xlsx_filename = f"{base_name}.xlsx"
        xlsx_path = os.path.join(batch_folder, xlsx_filename)
        create_xlsx_actes_societes(actes_data, xlsx_path, photo_name)
        generated_files.append(xlsx_filename)
        
    elif doc_type == "BIENS_IMMOBILIERS_ACHETEUR" or has_biens:
        # Only BIENS_IMMOBILIERS
        biens_data = extracted_data.copy()
        biens_data["doc_type"] = "BIENS_IMMOBILIERS_ACHETEUR"
        xlsx_filename = f"{base_name}.xlsx"
        xlsx_path = os.path.join(batch_folder, xlsx_filename)
        create_xlsx_biens_immobiliers(biens_data, xlsx_path, photo_name)
        generated_files.append(xlsx_filename)
        
    else:
        # Unknown type - create empty file with base name
        # Default to ACTES_SOCIETES structure
        mock_data = create_mock_data("ACTES_SOCIETES")
        xlsx_filename = f"{base_name}.xlsx"
        xlsx_path = os.path.join(batch_folder, xlsx_filename)
        create_xlsx_actes_societes(mock_data, xlsx_path, photo_name)
        generated_files.append(xlsx_filename)
    
    # ALWAYS generate IE_VISION_360 XLSX file (even if empty)
    # Try to extract BILAN data from the same image
    bilan_data = extract_bilan_with_openai(image_path)
    
    ie_filename = f"{base_name}__IE_VISION_360.xlsx"
    ie_path = os.path.join(batch_folder, ie_filename)
    create_ie_xlsx(extracted_data, ie_path, photo_name, bilan_data)
    generated_files.append(ie_filename)
    
    return generated_files


@app.route('/')
def index():
    """Render upload form."""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file upload and processing."""
    if 'files' not in request.files:
        flash('No files selected')
        return render_template('index.html')
    
    files = request.files.getlist('files')
    
    if not files or files[0].filename == '':
        flash('No files selected')
        return render_template('index.html')
    
    # Create batch folder
    batch_id = str(uuid.uuid4())
    batch_folder = os.path.join(app.config['OUTPUT_FOLDER'], batch_id)
    os.makedirs(batch_folder, exist_ok=True)
    
    results = []
    errors = []
    
    for file in files:
        if file and allowed_file(file.filename):
            try:
                # Secure filename
                filename = secure_filename(file.filename)
                
                # Save uploaded file temporarily
                temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(temp_path)
                
                # Process image
                generated_files = process_image(temp_path, filename, batch_folder)
                
                # Clean up temp file
                os.remove(temp_path)
                
                results.append({
                    'photo_name': filename,
                    'xlsx_files': generated_files,
                    'batch_id': batch_id
                })
                
            except Exception as e:
                errors.append({
                    'photo_name': file.filename,
                    'error': str(e)
                })
                print(f"Error processing {file.filename}: {str(e)}")
        else:
            errors.append({
                'photo_name': file.filename,
                'error': 'Invalid file type. Only JPG/PNG allowed.'
            })
    
    return render_template('results.html', results=results, errors=errors, batch_id=batch_id)


@app.route('/download/<batch_id>/<filename>')
def download_file(batch_id, filename):
    """Download individual XLSX file."""
    file_path = os.path.join(app.config['OUTPUT_FOLDER'], batch_id, filename)
    
    if not os.path.exists(file_path):
        flash('File not found')
        return render_template('index.html')
    
    return send_file(file_path, as_attachment=True, download_name=filename)


@app.route('/download_all/<batch_id>')
def download_all(batch_id):
    """Download all XLSX files as a ZIP archive."""
    batch_folder = os.path.join(app.config['OUTPUT_FOLDER'], batch_id)
    
    if not os.path.exists(batch_folder):
        flash('Batch not found')
        return render_template('index.html')
    
    # Create ZIP in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for root, dirs, files in os.walk(batch_folder):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = file  # Store files without folder structure
                zip_file.write(file_path, arcname)
    
    zip_buffer.seek(0)
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'batch_{batch_id}.zip'
    )


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
